import random
import re
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import process_survey123_field_data_classes as cls


global sssoc_info
global site_survey_info
global site_section_used
site_survey_info = []
sssoc_info = []  # site survey section observed collected
site_section_used = []

def read_in_excel_tab(wkbook_sheet):
    sheet = wkbook_sheet
    print('reading in {0}'.format(sheet.title))

    sheet_list = []
    i = 0

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column, values_only=True):
        sheet_list.append(row)
    

    return sheet_list

def read_in_excel_tab_header(wkbook_sheet):
    sheet = wkbook_sheet
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=sheet.max_column, values_only=True):
        return row
    
gear_types = {
    "1E Large": "EF_LB",
    "2E Large": "EF_LB",
    "3E Medium": "EF_MB",
    "4E Small": "EF_SB",
    "5E Medium Grassal": "EFG_MB",
    "6E Medium": "EF_MB",
    "7E Polycraft": "EF",
    "15E Large Grassal": "EFG_LB",
    "Punt 6": "Punt 6",
    "Punt 7": "Punt 7",
    "V Nose 9": "V Nose 9",
    "GT 10": "GT 10",
    "V Nose 11": "V Nose 11",
    "Canoe": "Canoe",
    "10ft Tinny": "10ft Tinny",
    "12ft Tinny": "12ft Tinny",
    "Bank Mounted": "EF_BM",
    "Back Pack": "EF_BP",
    "Unknown": "Unknown",
    "EXTRA_SHOT_IN_SAMPLES":"EXTRA_SHOT_IN_SAMPLES"
}

def get_random_shot(rs_site_id, rs_species):

    sp_col_found = 0

    # # Filter for shots with collected species at site
    rs_sub_sssoc_info = list(filter(lambda x: x[0] == rs_site_id and x[2] == rs_species and x[5] > 0, sssoc_info))
    # #    sub_sssoc_info = list(filter(lambda x: x__getitem__(0) == site_id and x__getitem__(2) == species and x__getitem__(5) > 0, sssoc_info))
    shotlist = []
    # print(str(len(rs_sub_sssoc_info)))
    if rs_sub_sssoc_info is None or len(rs_sub_sssoc_info) == 0:
        print('Notice: No collected {0} available in shots for site {1}'.format(rs_species, rs_site_id))

    if len(rs_sub_sssoc_info) > 0:
        # #        print('step 1')
        sp_col_found = 1
        prev_section_number = 0
        for rs_i in rs_sub_sssoc_info:
            if prev_section_number != rs_i.section_number:
                shotlist.append(rs_i.section_number)
                prev_section_number = rs_i.section_number

    if sp_col_found == 0:
        # # print('step 2')
        # # Filter for shots with species at site
        rs_sub_sssoc_info = list(filter(lambda x: x[0] == rs_site_id and x[2] == rs_species, sssoc_info))
        shotlist = []

        if len(rs_sub_sssoc_info) > 0:
            sp_col_found = 1
            prev_section_number = 0
            for rs_i in rs_sub_sssoc_info:
                if prev_section_number != rs_i.section_number:
                    shotlist.append(rs_i.section_number)
                    prev_section_number = rs_i.section_number

    if sp_col_found == 0:
        # print('step 3')
        # # Filter for shots at site
        # sub_sssoc_info = list(filter(lambda x: x[0] == rs_site_id and x[2] != 'No Fish', sssoc_info))
        sub_sssoc_info = list(filter(lambda x: x[0] == rs_site_id, sssoc_info))

        # print(rs_site_id)


        shotlist = []
        # print(len(sub_sssoc_info))
        if len(sub_sssoc_info) > 0:
            # sp_col_found = 1
            prev_section_number = 0
            for rs_i in sub_sssoc_info:
                if prev_section_number != rs_i.section_number:
                    shotlist.append(rs_i.section_number)
                    prev_section_number = rs_i.section_number
        # print(len(shotlist))

    if rs_sub_sssoc_info is None or len(shotlist) == 0:
        print('Notice: *** No {0}: {1} available'.format(rs_site_id, rs_species))
    else:
        return random.choice(shotlist)

def adjust_species_count(current, raw_data, PGID, section_num, species,svy_header, obs_header, sample_header, loc_header, shot_header, tally_results, tally_header):

    for completed in raw_data:
        #Check that site, section and species match:
        if PGID == completed.surveys[svy_header.index('GlobalID')]:
            if section_num == completed.shots[shot_header.index('section_number')] or section_num == completed.samples[sample_header.index('section_number_samp')]:
                if species == completed.observations[obs_header.index('species_obs')]:
                    #Adjust accordingly
                    if completed.observations[obs_header.index('section_collected')] is None:
                        collected_temp = 1
                    else:
                        collected_temp = completed.observations[obs_header.index('section_collected')]


                    if collected_temp > 1:
                        collected_temp -= (1 if current[sample_header.index('section_number_samp')] is None else current[sample_header.index('section_number_samp')])
            
                    else:
                        collected_temp = max(collected_temp, 0)
                        # collected_temp = collected_temp - 1
                        
                    completed.observations[obs_header.index('section_collected')] = collected_temp
                    
                    #Adjust Collected_Tally accordingly:
                    #Find tally data with the same PGID, section_num and species:
                    for tally in tally_results:
                       
                        if tally[tally_header.index('Site_ID')] == PGID:
                            
                            if tally[tally_header.index('Section_Number')] == section_num:
                                if tally[tally_header.index('Species')] == species:
                                    
                                    #Alter collected_tally:
                                    tally[tally_header.index('Collected_Tally')] = collected_temp


                    return
                

    return

def write_row(write_sheet, row_num: int, starting_column: str or int, write_values: list):
    if isinstance(starting_column, str):
        starting_column = ord(starting_column.lower()) - 96
    for wr_i, value in enumerate(write_values):
        write_sheet.cell(row_num, starting_column + wr_i, value)

def sheet_sort_rows(ws, row_start, row_end=0, cols=None, sorter=None, reverse=False):
    # #""" Sorts given rows of the sheet
    # #    row_start   First row to be sorted
    # #    row_end     Last row to be sorted (default last row)
    # #    cols        Columns to be considered in sort
    # #    sorter      Function that accepts a tuple of values and
    # #                returns a sortable key
    # #    reverse     Reverse the sort order
    # #"""

    bottom = ws.max_row
    if row_end == 0:
        row_end = ws.max_row
    right = openpyxl.utils.get_column_letter(ws.max_column)
    if cols is None:
        cols = range(1, ws.max_column+1)

    array = {}
    for ssr_row in range(row_start, row_end+1):
        key = []
        for col in cols:
            key.append(ws.cell(ssr_row, col).value)
        key = tuple(key)
        array[key] = array.get(key, set()).union({ssr_row})

    order = sorted(array, key=sorter, reverse=reverse)

    ws.move_range(f"A{row_start}:{right}{row_end}", bottom)
    dest = row_start
    for src_key in order:
        for ssr_row in array[src_key]:
            src = ssr_row + bottom
            dist = dest - src
            ws.move_range(f"A{src}:{right}{src}", dist)
            dest += 1

def set_col_date_style(ws, col_index):

    # create date style:
    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')

    for cds_row in ws[2:ws.max_row]:  # skip the header
        cell = cds_row[col_index]             # column H
        cell.style = date_style

def extra_record_output(ws, ero_site_id, ero_row_count):

    # ######### OUTPUT ANY EXTRA FISH CAUGHT BUT NOT MEASURED ######################################
    ## x[5] is collected_left count
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[5] > 0 and x[6] != 'IN SAMPLE INFO', sssoc_info))

    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'extra_caught')

    # ######### OUTPUT OBSERVED FISH ######################################
    ## x[4] is observed count
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[4] > 0, sssoc_info))

    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'observed')

    # ######### OUTPUT NO FISH ######################################
    ## x[2] is species name
    # #                if prev_sample_site_id == 'becd3e03-1cd0-44cc-8f3b-69cc65ef1957':
    # print('got here {0}'.format(ero_site_id))
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[2] == 'No Fish', sssoc_info))
    # print(len(sub_sssoc_info))
    # #                if prev_sample_site_id == 'becd3e03-1cd0-44cc-8f3b-69cc65ef1957' and len(sub_sssoc_info) <= 0:
    # #                    print('no hit for shot 8')
    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'no_fish')


    return ero_row_count

def extra_record_output_no_fish_shot(ws, ero_site_id, ero_section_number, ero_row_count):
    sub_sssoc_info = list(filter(lambda x: x[0] == ero_site_id and x[1] == str(ero_section_number), sssoc_info))
    ero_row_count = write_extra_data(ws, sub_sssoc_info, ero_row_count, 'no_shot_fish')

    return ero_row_count

def write_extra_data(ws_out, wed_sub_sssoc_info, r_count, extraDataType):

    if len(wed_sub_sssoc_info) > 0:
        for s in wed_sub_sssoc_info:

            wed_shot_i = s[1]
            if isinstance(wed_shot_i, str):
                # #                            print('converting shot')
                wed_shot_i = int(wed_shot_i)

            wed_sub_site_survey_info = list(filter(lambda x: x['k_site_id'] == s[0] and x['k_section_number'] == str(s[1]), site_survey_info))
            # print(s[0], s[1])
            if extraDataType == 'extra_caught':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING EXTRA CAUGHT for site: {0} shot: {1}'.format(s[0], s[1]))

            elif extraDataType == 'observed':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING OBSERVED for site: {0} shot: {1}'.format(s[0], s[1]))

            elif extraDataType == 'no_fish':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING NO FISH for site: {0} shot: {1}'.format(s[0], s[1]))

            elif extraDataType == 'no_shot_fish':
                if len(wed_sub_site_survey_info) > 1:
                    print('*** MULTIPLE SURVEY ERROR GETTING NO SHOT FISH for site: {0} shot: {1}'.format(s[0], s[1]))

                elif len(wed_sub_site_survey_info) == 0:
                    wed_sub_site_survey_info = list(filter(lambda x: x['k_site_id'] == s[0] and x['k_section_number'] == '1', site_survey_info))
                    if len(wed_sub_site_survey_info) > 0:
                        # wed_sub_site_survey_info[0]
                        write_excel_row(ws_out, r_count, wed_sub_site_survey_info[0], int(s[1]), 'No Fish', '', '', '', 0, 0, '', '', '', '', '', '', s[7], '')
                        print('*** ADDED EXTRA SHOT WITH NO FISH for site: {0} shot: {1}'.format(s[0], s[1]))
                        r_count += 1
                        return r_count
                    else:
                        print('*** NO SHOT 1 SURVEY INFO ERROR for site: {0} shot: 1'.format(s[0]))

            for wed_ss_row in wed_sub_site_survey_info:

                # #    0: self.site_id,
                # #    1: self.section_number,
                # #    2: self.species,
                # #    3: self.collected,
                # #    4: self.observed,
                # #    5: self.collected_left,
                # #    6: self.shot_id
                # #    7: self.obs_id
                if extraDataType == 'extra_caught':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, s[2], '', '', '', s[5], 0, '', '', '', '', '', '', s[7], '')
                    print('*** ADDED EXTRA CAUGHT for site: {0} shot: {1} species: {2}'.format(s[0], s[1], s[2]))

                elif extraDataType == 'observed':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, s[2], '', '', '', 0, s[4], '', '', '', '', '', '', s[7], '')
                    print('Notice: Added OBSERVED for site: {0} shot: {1} species: {2}'.format(s[0], s[1], s[2]))

                elif extraDataType == 'no_fish':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, s[2], '', '', '', 0, 0, '', '', '', '', '', '', s[7], '')
                    print('*** ADDED NO FISH for site: {0} shot: {1}'.format(s[0], s[1]))

                elif extraDataType == 'no_shot_fish':

                    write_excel_row(ws_out, r_count, wed_ss_row, wed_shot_i, 'No Fish', '', '', '', 0, 0, '', '', '', '', '', '', s[7], '')
                    print('*** ADDED NO FISH SHOT for site: {0} shot: {1}'.format(s[0], s[1]))

                r_count += 1
    return r_count

def write_excel_row(wsheet, rowcount, data_row, shot_num, wer_species, wer_fl, wer_tl, wer_w, wer_coll, wer_obs, wer_recapture, wer_pit, wer_external_tag_no, wer_genetics_label, wer_otoliths_label, wer_fauna_notes, wer_obst_id, wer_sample_id):

    if data_row['k_section_condition'].lower() == 'yes':
        section_condition_xl = 'FISHABLE'
    else:
        section_condition_xl = 'UNFISHABLE'

    personnel1 = data_row['k_personnel1']
    personnel2 = data_row['k_personnel2']

    #remove the common name within brackets
    wer_species = re.sub(r'\(.*?\) *', '', wer_species)
    wer_species = wer_species.strip()

    wer_gear_type = gear_types[data_row.gear_type]

##    wer_gear_type = data_row.gear_type #gear_types[data_row['k_gear_type']] if data_row['k_gear_type'] == data_row.gear_type else data_row.gear_type

    if data_row['k_survey_notes'] is None:
        wer_survey_notes = 'gear: {0}'.format(data_row['k_gear_type'])
    else:
        wer_survey_notes = '{0}, gear: {1}'.format(data_row['k_survey_notes'], data_row['k_gear_type'])

    wer_xl_row = list((data_row['k_project_name'], data_row['k_site_code'], data_row['k_x_start'], data_row['k_y_start'], data_row['k_x_finish'], data_row['k_y_finish'], data_row['k_survey_date'], wer_gear_type, personnel1, personnel2, data_row['k_depth_secchi'], data_row['k_depth_max'], data_row['k_depth_avg'], section_condition_xl, data_row['k_time_start'], data_row['k_time_end'], wer_survey_notes, shot_num, data_row['k_electro_seconds'], data_row['k_soak_minutes_per_unit'], data_row.section_time_start, data_row.section_time_end, data_row.volts, data_row.amps, data_row.pulses_per_second, data_row.percent_duty_cycle, wer_species, wer_fl, wer_tl, wer_w, wer_coll, wer_obs, wer_recapture, wer_pit, wer_external_tag_no, wer_genetics_label, wer_otoliths_label, wer_fauna_notes, data_row['k_water_qual_depth'], data_row['k_ec_25c'], data_row['k_water_temp'], data_row['k_do_mgl'], data_row['k_do_perc'], data_row['k_ph'], data_row['k_turbidity_ntu'], data_row['k_chlorophyll'], data_row['k_site_id'], data_row['k_shot_id'], wer_obst_id, wer_sample_id, data_row['k_data_x'], data_row['k_data_y']))

    write_row(wsheet, rowcount, 1, wer_xl_row)